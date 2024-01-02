"""
    aggregateOperationInstCount: extractvalue insertvalue
    arithmeticInstCount: add sub mul udiv sdiv urem srem icmp
    bitwiseOperationInstCount: shl lshr ashr and or xor
    callFunctionInstCount: call
    conversionInstCount: trunc zext sext fptrunc fpext fptoui fptosi uitofp sitofp ptrtoint inttoptr bitcast addrspacecast
    floatOperationInstCount: fadd fsub fmul fdiv frem fneg fcmp
    memoryAccessInstCount: alloca load store fence cmpxchg atomicrmw getelementptr
    terminatorInstCount: ret br switch indirectbr invoke callbr resume catchswitch catchret cleanupret unreachable
    vectorOperationInstCount: extractelement insertelement shufflevector
"""
agstrings_to_check = ["extractvalue", "insertvalue"]
arstrings_to_check = ["add", "sub", "mul", "udiv", "sdiv", "urem", "srem", "icmp"]
bstrings_to_check = ["shl", "lshr", "ashr", "and", "or", "xor"]
castrings_to_check = ["call"]
costrings_to_check = ["trunc", "zext", "sext", "fptrunc", "fpext", "fptoui", "fptosi", "uitofp", "sitofp", "ptrtoint", "inttoptr", "bitcast", "addrspacecast"]
fstrings_to_check = ["fadd", "fsub", "fmul", "fdiv", "frem", "fneg", "fcmp"]
mstrings_to_check = ["alloca", "load", "store", "fence", "cmpxchg", "atomicrmw", "getelementptr"]
tstrings_to_check = ["ret", "br", "switch", "indirectbr", "invoke", "callbr", "resume", "catchswitch", "catchret", "cleanupret", "unreachable"]
vstrings_to_check = ["extractelement", "insertelement", "shufflevector"]

import openpyxl

def check_strings(main_string, *strings_to_check):
    for string in strings_to_check:
        if string in main_string:
            return True
    return False

def count_instructions(file_path, benchmark_row0):
    aggregateOperationInstCount = 0
    arithmeticInstCount = 0
    bitwiseOperationInstCount = 0
    callFunctionInstCount = 0
    conversionInstCount = 0
    floatOperationInstCount = 0
    memoryAccessInstCount = 0
    terminatorInstCount = 0
    vectorOperationInstCount = 0

    # 打开.ll文件进行读取
    with open(file_path, 'r') as file:
        lines = file.readlines()

    # 遍历文件中的每一行
    for line in lines:
        if line.startswith(';'):
            continue
        
        if check_strings(line, *agstrings_to_check):
            aggregateOperationInstCount += 1

        if check_strings(line, *arstrings_to_check):
            arithmeticInstCount += 1

        if check_strings(line, *bstrings_to_check):
            bitwiseOperationInstCount += 1

        if check_strings(line, *castrings_to_check):
            callFunctionInstCount += 1

        if check_strings(line, *costrings_to_check):
            conversionInstCount += 1

        if check_strings(line, *fstrings_to_check):
            floatOperationInstCount += 1

        if check_strings(line, *mstrings_to_check):
           memoryAccessInstCount += 1

        if check_strings(line, *tstrings_to_check):
            terminatorInstCount += 1

        if check_strings(line, *vstrings_to_check):
            vectorOperationInstCount += 1


    # 打印统计结果
    print(f"聚合指令数量: {aggregateOperationInstCount}")
    print(f"算数指令数量: {arithmeticInstCount}")
    print(f"位运算指令数量: {bitwiseOperationInstCount}")
    print(f"调用指令数量: {callFunctionInstCount}")
    print(f"转换指令数量: {conversionInstCount}")
    print(f"浮点指令数量: {floatOperationInstCount}")
    print(f"访存指令数量: {memoryAccessInstCount}")
    print(f"终止指令数量: {terminatorInstCount}")
    print(f"向量指令数量: {vectorOperationInstCount}")

    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook.active
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row[0] and row[0].startswith(benchmark_row0):
            sheet.cell(row=row_number, column=2, value=aggregateOperationInstCount)
            sheet.cell(row=row_number, column=3, value=arithmeticInstCount)
            sheet.cell(row=row_number, column=4, value=bitwiseOperationInstCount)
            sheet.cell(row=row_number, column=5, value=callFunctionInstCount)
            sheet.cell(row=row_number, column=6, value=conversionInstCount)
            sheet.cell(row=row_number, column=7, value=floatOperationInstCount)
            sheet.cell(row=row_number, column=8, value=memoryAccessInstCount)
            sheet.cell(row=row_number, column=9, value=terminatorInstCount)
            sheet.cell(row=row_number, column=10, value=vectorOperationInstCount)
            break
    workbook.save('data.xlsx')


    
# 调用函数并传入.ll文件的路径
count_instructions('/home/zke/parboil-test/benchmarks/bfs/src/opencl_base/kernel.ll', 'parboil_bfs')
count_instructions('/home/zke/parboil-test/benchmarks/cutcp/src/opencl_base/kernel.ll', 'parboil_cutcp')
count_instructions('/home/zke/parboil-test/benchmarks/lbm/src/opencl_base/kernel.ll', 'parboil_lbm')
count_instructions('/home/zke/parboil-test/benchmarks/mri-q/src/opencl/kernels.ll', 'parboil_mri-q')
count_instructions('/home/zke/parboil-test/benchmarks/sgemm/src/opencl_base/kernel.ll', 'parboil_sgemm')
count_instructions('/home/zke/parboil-test/benchmarks/spmv/src/opencl_base/kernel.ll', 'parboil_spmv')
count_instructions('/home/zke/parboil-test/benchmarks/stencil/src/opencl_base/kernel.ll', 'parboil_stencil')
count_instructions('/home/zke/parboil-test/benchmarks/tpacf/src/opencl_base/kernel.ll', 'parboil_tpacf')

count_instructions('/home/zke/PolyBench-ACC/OpenCL/datamining/correlation/correlation.ll', 'polybench_correlation')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/datamining/covariance/covariance.ll', 'polybench_covariance')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/kernels/2mm/2mm.ll', 'polybench_2mm')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/kernels/3mm/3mm.ll', 'polybench_3mm')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/kernels/atax/atax.ll', 'polybench_atax')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/kernels/bicg/bicg.ll', 'polybench_bicg')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/kernels/doitgen/doitgen.ll', 'polybench_doitgen')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/kernels/gemm/gemm.ll', 'polybench_gemm')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/kernels/gemver/gemver.ll', 'polybench_gemver')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/kernels/gesummv/gesummv.ll', 'polybench_gesummv')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/kernels/mvt/mvt.ll', 'polybench_mvt')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/kernels/syr2k/syr2k.ll', 'polybench_syr2k')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/kernels/syrk/syrk.ll', 'polybench_syrk')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/solvers/gramschmidt/gramschmidt.ll', 'polybench_gramschmidt')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/linear-algebra/solvers/lu/lu.ll', 'polybench_lu')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/stencils/adi/adi.ll', 'polybench_adi')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/stencils/convolution-2d/2DConvolution.ll', 'polybench_convolution-2d')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/stencils/convolution-3d/3DConvolution.ll', 'polybench_convolution-3d')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/stencils/fdtd-2d/fdtd2d.ll', 'polybench_fdtd-2d')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/stencils/jacobi-1d-imper/jacobi1D.ll', 'polybench_jacobi-1d-imper')
count_instructions('/home/zke/PolyBench-ACC/OpenCL/stencils/jacobi-2d-imper/jacobi2D.ll', 'polybench_jacobi-2d-imper')

count_instructions('/home/zke/shoc/src/opencl/level1/fft/fft.ll', 'shoc_fft')
count_instructions('/home/zke/shoc/src/opencl/level1/gemm/gemmN.ll', 'shoc_gemm')
count_instructions('/home/zke/shoc/src/opencl/level1/md/md.ll', 'shoc_md')
count_instructions('/home/zke/shoc/src/opencl/level1/md5hash/md5.ll', 'shoc_md5hash')
count_instructions('/home/zke/shoc/src/opencl/level1/reduction/reduction.ll', 'shoc_reduction')
count_instructions('/home/zke/shoc/src/opencl/level1/scan/scan.ll', 'shoc_scan')
count_instructions('/home/zke/shoc/src/opencl/level1/sort/sort.ll', 'shoc_sort')
count_instructions('/home/zke/shoc/src/opencl/level1/spmv/spmv.ll', 'shoc_spmv')

count_instructions('/home/zke/rodinia-main/opencl/b+tree/kernel/kernel_gpu_opencl.ll', 'rodinia_b+tree')
count_instructions('/home/zke/rodinia-main/opencl/backprop/backprop_kernel.ll', 'rodinia_backprop')
count_instructions('/home/zke/rodinia-main/opencl/bfs/Kernels.ll', 'rodinia_bfs')
count_instructions('/home/zke/rodinia-main/opencl/cfd/Kernels.ll', 'rodinia_cfd')
count_instructions('/home/zke/rodinia-main/opencl/dwt2d/com_dwt.ll', 'rodinia_dwt2d')
count_instructions('/home/zke/rodinia-main/opencl/gaussian/gaussianElim_kernels.ll', 'rodinia_gaussian')
count_instructions('/home/zke/rodinia-main/opencl/heartwall/kernel/kernel_gpu_opencl.ll', 'rodinia_heartwall')
count_instructions('/home/zke/rodinia-main/opencl/hotspot/hotspot_kernel.ll', 'rodinia_hotspot')
count_instructions('/home/zke/rodinia-main/opencl/hotspot3D/hotspotKernel.ll', 'rodinia_hotspot3D')
count_instructions('/home/zke/rodinia-main/opencl/hybridsort/bucketsort_kernels.ll', 'rodinia_bucketsort')
count_instructions('/home/zke/rodinia-main/opencl/hybridsort/mergesort.ll', 'rodinia_mergesort')
count_instructions('/home/zke/rodinia-main/opencl/kmeans/kmeans.ll', 'rodinia_kmeans')
count_instructions('/home/zke/rodinia-main/opencl/lavaMD/kernel/kernel_gpu_opencl.ll', 'rodinia_lavaMD')
count_instructions('/home/zke/rodinia-main/opencl/lud/lud_kernel.ll', 'rodinia_lud')
count_instructions('/home/zke/rodinia-main/opencl/myocyte/kernel/kernel_gpu_opencl.ll', 'rodinia_myocyte')
count_instructions('/home/zke/rodinia-main/opencl/nn/nearestNeighbor_kernel.ll', 'rodinia_nn')
count_instructions('/home/zke/rodinia-main/opencl/nw/nw.cl', 'rodinia_nw')
count_instructions('/home/zke/rodinia-main/opencl/particlefilter/particle_double.ll', 'rodinia_particlefilter_double')
count_instructions('/home/zke/rodinia-main/opencl/particlefilter/particle_single.ll', 'rodinia_particlefilter_single')
count_instructions('/home/zke/rodinia-main/opencl/particlefilter/particle_naive.ll', 'rodinia_particlefilter_naive')
count_instructions('/home/zke/rodinia-main/opencl/pathfinder/kernels.ll', 'rodinia_pathfinder')
count_instructions('/home/zke/rodinia-main/opencl/srad/kernel/kernel_gpu_opencl.ll', 'rodinia_srad')
count_instructions('/home/zke/rodinia-main/opencl/streamcluster/Kernels.ll', 'rodinia_streamcluster')

count_instructions('/home/zke/AMDAPP_samples/cl/AtomicCounters/AtomicCounters_Kernels.ll', 'amd_AtomicCounters')
count_instructions('/home/zke/AMDAPP_samples/cl/BasicDebug/BasicDebug_Kernel.ll', 'amd_BasicDebug')
count_instructions('/home/zke/AMDAPP_samples/cl/BinarySearch/BinarySearch_Kernels.ll', 'amd_BinarySearch')
count_instructions('/home/zke/AMDAPP_samples/cl/BinomialOption/BinomialOption_Kernels.ll', 'amd_BinomialOption')
count_instructions('/home/zke/AMDAPP_samples/cl/BitonicSort/BitonicSort_Kernels.ll', 'amd_BitonicSort')
count_instructions('/home/zke/AMDAPP_samples/cl/BlackScholes/BlackScholes_Kernels.ll', 'amd_BlackScholes')
count_instructions('/home/zke/AMDAPP_samples/cl/BlackScholesDP/BlackScholesDP_Kernels.ll', 'amd_BlackScholesDP')
count_instructions('/home/zke/AMDAPP_samples/cl/BoxFilter/BoxFilter_Kernels.ll', 'amd_BoxFilter')
count_instructions('/home/zke/AMDAPP_samples/cl/BoxFilterGL/BoxFilterGL_Kernels.ll', 'amd_BoxFilterGL')
count_instructions('/home/zke/AMDAPP_samples/cl/BufferBandwidth/BufferBandwidth_Kernels.ll', 'amd_BufferBandwidth')
count_instructions('/home/zke/AMDAPP_samples/cl/BufferImageInterop/BufferImageInterop_kernels.ll', 'amd_BufferImageInterop')
count_instructions('/home/zke/AMDAPP_samples/cl/DCT/DCT_Kernels.ll', 'amd_DCT')
count_instructions('/home/zke/AMDAPP_samples/cl/DeviceFission/DeviceFission_Kernels.ll', 'amd_DeviceFission')
count_instructions('/home/zke/AMDAPP_samples/cl/DeviceFission11Ext/DeviceFission11Ext_Kernels.ll', 'amd_DeviceFission11Ext')
count_instructions('/home/zke/AMDAPP_samples/cl/DwtHaar1D/DwtHaar1D_Kernels.ll', 'amd_DwtHaar1D')
count_instructions('/home/zke/AMDAPP_samples/cl/FastWalshTransform/FastWalshTransform_Kernels.ll', 'amd_FastWalshTransform')
count_instructions('/home/zke/AMDAPP_samples/cl/FloydWarshall/FloydWarshall_Kernels.ll', 'amd_FloydWarshal')
count_instructions('/home/zke/AMDAPP_samples/cl/FluidSimulation2D/FluidSimulation2D_Kernels.ll', 'amd_FluidSimulation2D')
count_instructions('/home/zke/AMDAPP_samples/cl/GaussianNoiseGL/GaussianNoiseGL_Kernels.ll', 'amd_GaussianNoiseGL')
count_instructions('/home/zke/AMDAPP_samples/cl/HelloWorld/HelloWorld_Kernel.ll', 'amd_HelloWorld')
count_instructions('/home/zke/AMDAPP_samples/cl/Histogram/Histogram_Kernels.ll', 'amd_Histogram')
count_instructions('/home/zke/AMDAPP_samples/cl/ImageBandwidth/ImageBandwidth_Kernels.ll', 'amd_ImageBandwidth')
count_instructions('/home/zke/AMDAPP_samples/cl/ImageOverlap/ImageOverlap_Kernels.ll', 'amd_ImageOverlap')
count_instructions('/home/zke/AMDAPP_samples/cl/KernelLaunch/KernelLaunch_Kernels.ll', 'amd_KernelLaunch')
count_instructions('/home/zke/AMDAPP_samples/cl/KmeansAutoclustering/KmeansAutoclustering_Kernels.ll', 'amd_KmeansAutoclustering')
count_instructions('/home/zke/AMDAPP_samples/cl/LUDecomposition/LUDecomposition_Kernels.ll', 'amd_LUDecomposition')
count_instructions('/home/zke/AMDAPP_samples/cl/Mandelbrot/Mandelbrot_Kernels.ll', 'amd_Mandelbrot')
count_instructions('/home/zke/AMDAPP_samples/cl/MatrixMulImage/MatrixMulImage_Kernels.ll', 'amd_MatrixMulImage')
count_instructions('/home/zke/AMDAPP_samples/cl/MatrixMultiplication/MatrixMultiplication_Kernels.ll', 'amd_MatrixMultiplication')
count_instructions('/home/zke/AMDAPP_samples/cl/MatrixTranspose/MatrixTranspose_Kernels.ll', 'amd_MatrixTranspose')
count_instructions('/home/zke/AMDAPP_samples/cl/MemoryModel/MemoryModel.ll', 'amd_MemoryModel')
count_instructions('/home/zke/AMDAPP_samples/cl/MonteCarloAsian/MonteCarloAsian_Kernels.ll', 'amd_MonteCarloAsian')
count_instructions('/home/zke/AMDAPP_samples/cl/MonteCarloAsianDP/MonteCarloAsianDP_Kernels.ll', 'amd_MonteCarloAsianDP')
count_instructions('/home/zke/AMDAPP_samples/cl/NBody/NBody_Kernels.ll', 'amd_NBody')
count_instructions('/home/zke/AMDAPP_samples/cl/PrefixSum/PrefixSum_Kernels.ll', 'amd_PrefixSum')
count_instructions('/home/zke/AMDAPP_samples/cl/QuasiRandomSequence/QuasiRandomSequence_Kernels.ll', 'amd_QuasiRandomSequence')
count_instructions('/home/zke/AMDAPP_samples/cl/RadixSort/RadixSort_Kernels.ll', 'amd_RadixSort')
count_instructions('/home/zke/AMDAPP_samples/cl/RecursiveGaussian/RecursiveGaussian_Kernels.ll', 'amd_RecursiveGaussian')
count_instructions('/home/zke/AMDAPP_samples/cl/Reduction/Reduction_Kernels.ll', 'amd_Reduction')
count_instructions('/home/zke/AMDAPP_samples/cl/ScanLargeArrays/ScanLargeArrays_Kernels.ll', 'amd_ScanLargeArrays')
count_instructions('/home/zke/AMDAPP_samples/cl/SimpleConvolution/SimpleConvolution_Kernels.ll', 'amd_SimpleConvolution')
count_instructions('/home/zke/AMDAPP_samples/cl/SimpleGL/SimpleGL_Kernels.ll', 'amd_SimpleGL')
count_instructions('/home/zke/AMDAPP_samples/cl/SimpleImage/SimpleImage_Kernels.ll', 'amd_SimpleImage')
count_instructions('/home/zke/AMDAPP_samples/cl/SobelFilter/SobelFilter_Kernels.ll', 'amd_SobelFilter')
count_instructions('/home/zke/AMDAPP_samples/cl/StringSearch/StringSearch_Kernels.ll', 'amd_StringSearch')
count_instructions('/home/zke/AMDAPP_samples/cl/Template/Template_Kernels.ll', 'amd_Template')
count_instructions('/home/zke/AMDAPP_samples/cl/TransferOverlap/TransferOverlap_Kernels.ll', 'amd_TransferOverlap')
count_instructions('/home/zke/AMDAPP_samples/cl/URNG/URNG_Kernels.ll', 'amd_URNG')
