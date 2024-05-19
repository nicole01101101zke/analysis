import os
import subprocess
import re
import openpyxl

data_path = "/root/pocl-static-analysis/regressionModel/data.xlsx" # file to store analysis result
parboil_directory = "/root/parboil_CPU"
rodinia_directory = "/root/rodinia_CPU/opencl"
rodinia_particlefilter_directory = "/root/rodinia_CPU/opencl/particlefilter"
amd_directory = "/root/AMDAPP_samples_CPU/cl"

def run_parboil(benchmark_row0, benchmark, version, dataset):
    os.chdir(parboil_directory)
    compile_command = f"./parboil run {benchmark} {version} {dataset}"
    result = subprocess.run(compile_command, shell=True, stdout=subprocess.PIPE, text=True)
    output = result.stdout
    match = re.search(r'Timer Wall Time:\s+(\d+\.\d+)', output)

    if match:
        extracted_number = float(match.group(1))
        print("Extracted Number:", extracted_number)
    else:
        print("No match found.")
    formatted_number = "{:.10f}".format(extracted_number)
    
    workbook = openpyxl.load_workbook(data_path)
    sheet = workbook.active
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number > 232 and row[0] and row[0].startswith(benchmark_row0):
            sheet.cell(row=row_number, column=23, value=formatted_number)
            break
    workbook.save(data_path)

    
def run_rodinia(benchmark_row0, path):
    os.chdir(rodinia_directory)
    compile_command = f"cd {path} && make clean && make && make --always-make run" # make rundwt2d报错make: 'run' is up to date.
    result = subprocess.run(compile_command, shell=True, stdout=subprocess.PIPE, text=True)
    output = result.stdout
    match = re.search(r'Total:\s*([\d.]+)', output)
    match_leu = re.search(r'Total application run time:\s*([\d.]+)', output)

    if match_leu:
        extracted_number = float(match_leu.group(1))
        print("Extracted Number:", extracted_number)
    elif match:
        extracted_number = float(match.group(1))
        print("Extracted Number:", extracted_number)
    else:
        print("No match found.")
    formatted_number = "{:.10f}".format(extracted_number)
    
    workbook = openpyxl.load_workbook(data_path)
    sheet = workbook.active
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number > 232 and row[0] and row[0].startswith(benchmark_row0):
            sheet.cell(row=row_number, column=23, value=formatted_number)
            break
    workbook.save(data_path)

def run_rodinia_nn(): # 这个看的是Exec:
    formatted_number = "{:.10f}".format(0.090730)
    
    workbook = openpyxl.load_workbook(data_path)
    sheet = workbook.active
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number > 232 and row[0] and row[0].startswith("rodinia_nn"):
            sheet.cell(row=row_number, column=23, value=formatted_number)
            break
    workbook.save(data_path)

def run_rodinia_pathfinder(): # 很奇怪，编译时找不到timing.h，在这里跑找不到编译好的文件
    formatted_number = "{:.10f}".format(1.191245) 
    
    workbook = openpyxl.load_workbook(data_path)
    sheet = workbook.active
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number > 232 and row[0] and row[0].startswith("rodinia_pathfinder"):
            sheet.cell(row=row_number, column=23, value=formatted_number)
            break
    workbook.save(data_path)

def run_rodinia_particlefilter(benchmark_row0, command):
    os.chdir(rodinia_particlefilter_directory)
    compile_command = f"make clean && make && {command}"
    result = subprocess.run(compile_command, shell=True, stdout=subprocess.PIPE, text=True)
    output = result.stdout
    match = re.search(r'Total:\s*([\d.]+)', output)

    if match:
        extracted_number = float(match.group(1))
        print("Extracted Number:", extracted_number)
    else:
        print("No match found.")
    formatted_number = "{:.10f}".format(extracted_number)
    
    workbook = openpyxl.load_workbook(data_path)
    sheet = workbook.active
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number > 232 and row[0] and row[0].startswith(benchmark_row0):
            sheet.cell(row=row_number, column=23, value=formatted_number)
            break
    workbook.save(data_path)

def run_AMD(benchmark_row0, path, run_command):
    os.chdir(amd_directory)
    compile_command = f"cd {path} && cd build/bin/x86_64/Release/cl/{path} && {run_command}" 
    result = subprocess.run(compile_command, shell=True, stdout=subprocess.PIPE, text=True)
    output = result.stdout
    match = re.search(r'Total:\s*([\d.]+)', output)

    if match:
        extracted_number = float(match.group(1))
        print("Extracted Number:", extracted_number)
    else:
        print("No match found.")
    formatted_number = "{:.10f}".format(extracted_number)
    
    workbook = openpyxl.load_workbook(data_path)
    sheet = workbook.active
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number > 232 and row[0] and row[0].startswith(benchmark_row0):
            sheet.cell(row=row_number, column=23, value=formatted_number)
            break
    workbook.save(data_path)


# run_parboil("parboil_cutcp", "cutcp", "opencl_base", "large")
# run_parboil("parboil_lbm", "lbm", "opencl_base", "long")
# run_parboil("parboil_mri-q", "mri-q", "opencl", "large")
# run_parboil("parboil_stencil", "stencil", "opencl_base", "default")
# run_parboil("parboil_tpacf", "tpacf", "opencl_base", "large")

# run_rodinia("rodinia_b+tree", "b+tree")
# run_rodinia("rodinia_backprop", "backprop")
# run_rodinia("rodinia_bfs", "bfs")
# run_rodinia("rodinia_cfd", "cfd") # run操作没写在makefile里，修改了Makefile
# run_rodinia("rodinia_dwt2d", "dwt2d") # run操作没写在makefile里，修改了Makefile
# run_rodinia("rodinia_gaussian", "gaussian")
# run_rodinia("rodinia_heartwall", "heartwall") # linker找不到main.h，把main.h的内容写在了kernel里
# run_rodinia("rodinia_hotspot", "hotspot")
# run_rodinia("rodinia_hotspot3D", "hotspot3D")
# run_rodinia("rodinia_hybridsort", "hybridsort")
# run_rodinia("rodinia_kmeans", "kmeans") # Segmentation fault (core dumped) 暂时不处理
# run_rodinia("rodinia_lavaMD", "lavaMD")
# run_rodinia("rodinia_leukocyte", "leukocyte") # Error: None of the platforms has a GPU
# run_rodinia("rodinia_lud", "lud")
# run_rodinia("rodinia_myocyte", "myocyte") #这个的时间消耗很清楚，包括了GPU传输数据的时间
run_rodinia_nn() # 这个有交互，需要Enter Platform and Device No (Seperated by Space)
# run_rodinia("rodinia_nw", "nw")
# run_rodinia_pathfinder()
# run_rodinia("rodinia_srad", "srad") # linker找不到main.h，把main.h的内容写在了kernel里
# run_rodinia("rodinia_streamcluster", "streamcluster")

# ERROR: clGetContextInfo() => CL_INVALID_VALUE
# ERROR: clGetProgramBuildInfo() => -33
# run_rodinia_particlefilter("rodinia_particlefilter_naive", "./OCL_particlefilter_naive.out -x 128 -y 128 -z 10 -np 10000")

# run_rodinia_particlefilter("rodinia_particlefilter_single","./OCL_particlefilter_single.out -x 128 -y 128 -z 10 -np 400000 $@")
# run_rodinia_particlefilter("rodinia_particlefilter_double","./OCL_particlefilter_double.out -x 128 -y 128 -z 10 -np 400000 $@")

# run_AMD("amd_AtomicCounters", "AtomicCounters", "./AtomicCounters -t -x 16777216 -i 100")
# run_AMD("amd_BinarySearch", "BinarySearch", "./BinarySearch -t -i 100")
# run_AMD("amd_BinomialOption", "BinomialOption", "./BinomialOption -t -i 100")
# run_AMD("amd_BitonicSort", "BitonicSort", "./BitonicSort -t -i 100")
# run_AMD("amd_BlackScholes", "BlackScholes", "./BlackScholes -t -i 100")
# run_AMD("amd_BlackScholesDP", "BlackScholesDP", "./BlackScholesDP -t -i 100")
# run_AMD("amd_BoxFilter", "BoxFilter", "./BoxFilter -t -i 100")
# # run_AMD("amd_BufferBandwidth", "BufferBandwidth", "./BufferBandwidth -t -i 100") #这个专门用于测读写带宽，没必要放进来
# run_AMD("amd_BufferImageInterop", "BufferImageInterop", "./BufferImageInterop -t -i 100")

# run_AMD("amd_DCT", "DCT", "./DCT -t -i 100")
# # run_AMD("amd_DeviceFission", "DeviceFission", "./DeviceFission -t -i 100") # 设备不支持
# # run_AMD("amd_DeviceFission11Ext", "DeviceFission11Ext", "./DeviceFission11Ext -t -i 100") # 设备不支持
# run_AMD("amd_DwtHaar1D", "DwtHaar1D", "./DwtHaar1D -t -i 100")
# run_AMD("amd_FastWalshTransform", "FastWalshTransform", "./FastWalshTransform -t -i 100")
# run_AMD("amd_FloydWarshall", "FloydWarshall", "./FloydWarshall -t -i 100")
# # run_AMD("amd_FluidSimulation2D", "FluidSimulation2D", "./FluidSimulation2D -t -i 100") # 依赖缺失，不跑
# # run_AMD("amd_GaussianNoiseGL", "GaussianNoiseGL", "./GaussianNoiseGL -t -i 100") # 依赖缺失，不跑

# # Error: clGetEventEventInfo Failed with Error Code: Error code : unknown error code
# # Location : /home/zke/AMDAPP_samples_GPU/cl/Histogram/../../include/SDKUtil/CLUtil.hpp:830
# # Error: WaitForEventAndRelease(ndrEvt1) Failed
# # run_AMD("amd_Histogram", "Histogram", "./Histogram -t -i 100") # error
# # run_AMD("amd_ImageBandwidth", "ImageBandwidth", "./ImageBandwidth -t -i 100") #专门用于测读写带宽，没必要放进来
# run_AMD("amd_ImageOverlap", "ImageOverlap", "./ImageOverlap -t -i 100")
# # run_AMD("amd_KernelLaunch", "KernelLaunch", "./KernelLaunch -t -i 100") #专门用于测读写带宽，没必要放进来
# # run_AMD("amd_KmeansAutoclustering", "KmeansAutoclustering", "./KmeansAutoclustering -t -i 100") # 依赖缺失，不跑
# run_AMD("amd_LUDecomposition", "LUDecomposition", "./LUDecomposition -t -i 100")
# # run_AMD("amd_Mandelbrot", "Mandelbrot", "./Mandelbrot -t -i 100") # 依赖缺失，不跑
# run_AMD("amd_MatrixMulImage", "MatrixMulImage", "./MatrixMulImage -t -i 100")
# run_AMD("amd_MatrixMultiplication", "MatrixMultiplication", "./MatrixMultiplication -t --eAppGflops")
# run_AMD("amd_MatrixTranspose", "MatrixTranspose", "./MatrixTranspose -t -i 100")
# # run_AMD("amd_MemoryModel", "MemoryModel", "./MemoryModel -t -i 100") # 不统计时间
# run_AMD("amd_MonteCarloAsian", "MonteCarloAsian", "./MonteCarloAsian -t -i 100")
# run_AMD("amd_MonteCarloAsianDP", "MonteCarloAsianDP", "./MonteCarloAsianDP -t -i 100")

# # run_AMD("amd_NBody", "NBody", "./NBody -t -i 100") # 依赖缺失，不跑
# run_AMD("amd_PrefixSum", "PrefixSum", "./PrefixSum -t -i 100")
# run_AMD("amd_QuasiRandomSequence", "QuasiRandomSequence", "./QuasiRandomSequence -t -i 100")
# run_AMD("amd_RadixSort", "RadixSort", "./RadixSort -t -i 100")
# run_AMD("amd_RecursiveGaussian", "RecursiveGaussian", "./RecursiveGaussian -t -i 100")
# run_AMD("amd_Reduction", "Reduction", "./Reduction -t -i 100")
# run_AMD("amd_ScanLargeArrays", "ScanLargeArrays", "./ScanLargeArrays -t -i 100")
# run_AMD("amd_SimpleConvolution", "SimpleConvolution", "./SimpleConvolution -t -i 100")
# # run_AMD("amd_SimpleGL", "SimpleGL", "./SimpleGL -t -i 100") # 依赖缺失，不跑
# run_AMD("amd_SimpleImage", "SimpleImage", "./SimpleImage -t -i 100")
# run_AMD("amd_SobelFilter", "SobelFilter", "./SobelFilter -t -i 100")
# run_AMD("amd_StringSearch", "StringSearch", "./StringSearch -t -i 100")
# # run_AMD("amd_Template", "Template", "./Template -t -i 100") # 不统计时间
# run_AMD("amd_TransferOverlap", "TransferOverlap", "./TransferOverlap -t -i 100")
# run_AMD("amd_URNG", "URNG", "./URNG -t -i 100")
